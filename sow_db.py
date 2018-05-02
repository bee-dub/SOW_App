import sqlite3
from openpyxl import load_workbook


class SOW_Database:
    def __init__(self):
        self.conn = sqlite3.connect('sow_database.db')
        self.cursor = self.conn.cursor()

    def __del__(self):
        self.conn.close()

    def create_table(self):
        sql = \
            '''
            CREATE TABLE sow(S_No integer, BU_Code text, SOW_ID text, CHC_BU text, Type text, Tower text,
                SOW_Name text, Engagement_Model text, SOW_Owner_Wipro text, SOW_Owner_CHC text, Offshore integer,
                Onsite integer, Total_FTE real, SOW_Value integer, Start_Date date, End_Date date, Status text, Remarks text)
            '''
        self.cursor.execute(sql)
        self.conn.commit()

    def import_sow_data(self):
        qry = \
            '''
            insert into sow(S_No,BU_Code,SOW_ID,CHC_BU,Type,Tower,SOW_Name,Engagement_Model,SOW_Owner_Wipro,
                SOW_Owner_CHC,Offshore,Onsite,Total_FTE,SOW_Value,Start_Date,End_Date,Status,Remarks) 
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            '''
        wb = load_workbook(r"C:\Users\esizzsx\Desktop\sow.xlsx", read_only=True)
        ws = wb['Sheet1']
        for row in ws.iter_rows(min_row=2):
            s_no = row[0].value
            bu_code = row[1].value
            sow_id = row[2].value
            chc_bu = row[3].value
            type = row[4].value
            tower = row[5].value
            sow_name = row[6].value
            engagement_model = row[7].value
            sow_owner_wipro = row[8].value
            sow_owner_chc = row[9].value
            offshore = row[10].value
            onsite = row[11].value
            total_fte = row[12].value
            sow_value = row[13].value
            start_date = row[14].value
            end_date = row[15].value
            status = row[16].value
            remarks = row[17].value

            values = (s_no, bu_code, sow_id, chc_bu, type, tower, sow_name, engagement_model, sow_owner_wipro,
                      sow_owner_chc, offshore, onsite, total_fte, sow_value, start_date, end_date, status, remarks)
            self.cursor.execute(qry, values)
        self.conn.commit()

    def sow_records(self):
        self.cursor.execute('SELECT * FROM sow')
        records = self.cursor.fetchall()
        return records

    def delete_record(self, Id):
        self.cursor.execute('DELETE FROM sow WHERE S_No=?', (Id, ))
        self.conn.commit()

    def insert_record(self, *args):
        sql = 'INSERT INTO sow VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'
        self.cursor.execute(sql, *args)
        self.conn.commit()

    def update_record(self, *args):
        sql = \
            '''
            UPDATE sow SET BU_Code=?, SOW_ID=?, CHC_BU=?, Type=?, Tower=?, SOW_Name=?, Engagement_Model=?,
                SOW_Owner_Wipro=?, SOW_Owner_CHC=?, Offshore=?, Onsite=?, Total_FTE=?, SOW_Value=?, Start_Date=?, 
                End_Date=?, Status=?, Remarks=? 
            WHERE S_No=? 
            '''
        self.cursor.execute(sql, *args)
        self.conn.commit()

    def search_records(self, sow_name='', sow_owner_wipro='', sow_owner_chc=''):
        if sow_name != '' or sow_owner_wipro != '' or sow_owner_chc != '':
            sow_name = '*****' if sow_name == '' else sow_name
            sow_owner_wipro = '*****' if sow_owner_wipro == '' else sow_owner_wipro
            sow_owner_chc = '*****' if sow_owner_chc == '' else sow_owner_chc
        self.cursor.execute('SELECT * FROM sow WHERE SOW_Name LIKE ? OR SOW_Owner_Wipro LIKE ? OR SOW_Owner_CHC LIKE ?',
                            ('%{}%'.format(sow_name), '%{}%'.format(sow_owner_wipro), '%{}%'.format(sow_owner_chc), ))
        records = self.cursor.fetchall()
        return records




